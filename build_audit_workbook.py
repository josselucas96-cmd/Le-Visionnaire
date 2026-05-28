"""Generate `daily_holdings_audit.xlsx` — the canonical audit + recovery document.

Run after every move (or daily) to refresh:
    python build_audit_workbook.py

What it produces (7 sheets):
- README                : overview + per-portfolio summary
- <Portfolio>           : daily_holdings pivot (date × ticker, with shares & price)
                           Last row = today, populated with live yfinance prices.
- <Portfolio> Moves     : event log + inception state, enough to fully
                           reconstruct the portfolio from scratch.

Source of truth = Supabase. yfinance only for today's prices.

This document is the AUDIT + RECOVERY artifact: if Supabase loses data, the
workbook + the snapshots/<portfolio>/<inception>.csv files are enough to
replay every move and rebuild the current state.
"""
import csv
from collections import defaultdict
from datetime import date as _date, datetime
from pathlib import Path

import toml
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import create_client


# Read portfolio list dynamically from DB (same pattern as daily_refresh.py)
# This auto-includes new portfolios (e.g., test sandbox) without code change.
def _get_all_portfolio_ids(sb):
    rows = sb.table("portfolios").select("id").order("display_order").execute().data
    return [r["id"] for r in rows]
TODAY = _date.today().isoformat()
WORKBOOK_PATH = Path(__file__).parent / "daily_holdings_audit.xlsx"

THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", start_color="22305A", end_color="22305A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TODAY_FILL  = PatternFill("solid", start_color="FFF7DC", end_color="FFF7DC")
SECTION_FILL = PatternFill("solid", start_color="EEEEEE", end_color="EEEEEE")
SECTION_FONT = Font(bold=True, size=11)


# ── yfinance ──────────────────────────────────────────────────────────────────
_PRICE_CACHE: dict[str, float | None] = {}

# FX (standalone copy of utils.market.usd_factor — no streamlit in the cron).
# Foreign listings are converted to USD; minor units (pence GBp) divide by 100.
_MINOR_UNIT = {"GBp": "GBP", "GBX": "GBP", "ZAc": "ZAR", "ZAX": "ZAR", "ILA": "ILS"}
_FX_CACHE: dict[str, float | None] = {"USD": 1.0}

def _usd_factor(ccy: str | None) -> float:
    if not ccy or ccy == "USD":
        return 1.0
    if ccy not in _FX_CACHE:
        major = _MINOR_UNIT.get(ccy, ccy)
        try:
            _FX_CACHE[ccy] = float(yf.Ticker(f"{major}USD=X").fast_info.last_price)
        except Exception:
            _FX_CACHE[ccy] = None
    r = _FX_CACHE[ccy]
    if r is None:
        return 1.0  # unknown rate -> leave native (avoid crash)
    return r / 100.0 if ccy in _MINOR_UNIT else r

def fetch_price(ticker: str) -> float | None:
    """Latest price in USD (converts foreign listings)."""
    if ticker == "CASH":
        return 1.0
    if ticker in _PRICE_CACHE:
        return _PRICE_CACHE[ticker]
    try:
        fi = yf.Ticker(ticker).fast_info
        p = float(fi.last_price) * _usd_factor(getattr(fi, "currency", None))
        _PRICE_CACHE[ticker] = p
        return p
    except Exception as e:
        print(f"  ! {ticker}: yfinance fetch failed ({e})")
        _PRICE_CACHE[ticker] = None
        return None


# ── Holdings sheet ────────────────────────────────────────────────────────────
def fetch_holdings(sb, portfolio_id: str) -> list[dict]:
    """Paginated fetch — Supabase default limit is 1000 rows. Portfolio_Test
    has 3700+ rows (138 days × 27 tickers); without pagination the last
    date would be partial and the workbook NAV calc wrong."""
    rows = []
    offset = 0
    PAGE = 1000
    while True:
        chunk = (
            sb.table("daily_holdings")
            .select("date, ticker, shares, price, value")
            .eq("portfolio_id", portfolio_id)
            .order("date")
            .order("ticker")
            .range(offset, offset + PAGE - 1)
            .execute()
            .data
        )
        if not chunk:
            break
        rows.extend(chunk)
        if len(chunk) < PAGE:
            break
        offset += PAGE
    return rows


def build_holdings_sheet(wb: Workbook, portfolio_id: str, rows: list[dict],
                         active_positions: list[dict]) -> tuple[str, float]:
    """Build the daily_holdings pivot sheet. Append a row for today with live prices.

    Returns (sheet_name, nav_today) for the README summary.
    """
    sheet_name = portfolio_id.capitalize()
    ws = wb.create_sheet(sheet_name)

    # Group by date
    by_date: dict[str, dict[str, dict]] = defaultdict(dict)
    tickers_seen: set[str] = set()
    for r in rows:
        by_date[r["date"]][r["ticker"]] = r
        tickers_seen.add(r["ticker"])

    # CASH first then equities alphabetically
    equity_tickers = sorted(t for t in tickers_seen if t != "CASH")
    ordered_tickers = (["CASH"] if "CASH" in tickers_seen else []) + equity_tickers

    # Header
    header = ["Date", "NAV Total ($)"]
    for tk in ordered_tickers:
        header.extend([f"{tk}_value", f"{tk}_shares", f"{tk}_price"])
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    # Historical rows (literal values from DB)
    for d in sorted(by_date.keys()):
        row = [d]
        nav_total = sum(float(r["value"] or 0) for r in by_date[d].values())
        row.append(round(nav_total, 2))
        for tk in ordered_tickers:
            r = by_date[d].get(tk)
            if r is None:
                row.extend([None, None, None])
            else:
                row.extend([
                    round(float(r["value"] or 0), 2),
                    round(float(r["shares"] or 0), 8),
                    round(float(r["price"] or 0), 4),
                ])
        ws.append(row)

    # Today's row — fetch live prices, reuse last row's shares (no moves today
    # are reflected in daily_holdings until visitor traffic triggers
    # lazy_write_holdings). The cash $ is derived from get_cash_amount logic
    # inline: latest CASH value strictly before today + today's tx deltas.
    today_row_idx = ws.max_row + 1
    last_data_row = ws.max_row
    today_cash = compute_today_cash(rows, active_positions, portfolio_id)

    today_values: list = [TODAY]
    # NAV total: formula sum of *_value columns
    value_col_indices = [3 + i * 3 for i in range(len(ordered_tickers))]
    nav_formula = "=" + "+".join(
        f"{get_column_letter(c)}{today_row_idx}" for c in value_col_indices
    )
    today_values.append(nav_formula)

    for i, tk in enumerate(ordered_tickers):
        if tk == "CASH":
            today_values.extend([
                round(today_cash, 2),
                round(today_cash, 2),
                1.0,
            ])
        else:
            shares = active_shares_for(active_positions, tk, fallback_row=by_date[max(by_date)].get(tk))
            price = fetch_price(tk)
            if price is None:
                # Fallback to last known price
                last_row = by_date[max(by_date)].get(tk)
                price = float(last_row["price"]) if last_row else 0.0
            # Value as formula = shares × price (so it stays editable)
            scol = get_column_letter(3 + i * 3 + 1)
            pcol = get_column_letter(3 + i * 3 + 2)
            today_values.extend([
                f"={scol}{today_row_idx}*{pcol}{today_row_idx}",
                round(shares, 8),
                round(price, 4),
            ])

    ws.append(today_values)
    # Highlight today's row
    for c in range(1, len(header) + 1):
        ws.cell(row=today_row_idx, column=c).fill = TODAY_FILL
    ws.cell(row=today_row_idx, column=1).font = Font(bold=True)
    ws.cell(row=today_row_idx, column=2).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    for i, _tk in enumerate(ordered_tickers):
        base_col = 3 + i * 3
        ws.column_dimensions[get_column_letter(base_col)].width = 14
        ws.column_dimensions[get_column_letter(base_col + 1)].width = 14
        ws.column_dimensions[get_column_letter(base_col + 2)].width = 11

    ws.freeze_panes = "C2"

    # Approximate NAV today (sum of value cells, but those are formulas — compute
    # from raw numbers we just wrote)
    nav_today = today_cash
    for i, tk in enumerate(ordered_tickers):
        if tk == "CASH":
            continue
        shares = active_shares_for(active_positions, tk, fallback_row=by_date[max(by_date)].get(tk))
        price = _PRICE_CACHE.get(tk)
        if price is None:
            last_row = by_date[max(by_date)].get(tk)
            price = float(last_row["price"]) if last_row else 0.0
        nav_today += shares * price

    return sheet_name, nav_today


def active_shares_for(active_positions: list[dict], ticker: str,
                       fallback_row: dict | None) -> float:
    for p in active_positions:
        if p["ticker"] == ticker:
            return float(p.get("shares") or 0)
    # Fallback: shares from yesterday's daily_holdings row
    return float(fallback_row["shares"]) if fallback_row else 0.0


def compute_today_cash(rows: list[dict], active_positions: list[dict],
                       portfolio_id: str) -> float:
    """Derive today's cash $ — same logic as utils/data.get_cash_amount."""
    # Latest CASH row strictly before today
    cash_rows = [r for r in rows if r["ticker"] == "CASH" and r["date"] < TODAY]
    if cash_rows:
        cash_rows.sort(key=lambda r: r["date"], reverse=True)
        baseline = float(cash_rows[0]["value"])
    else:
        baseline = 0.0
    # Today's transaction deltas (queried inline in main loop, passed in via
    # active_positions? actually we re-query below from `_TODAY_TXNS_CACHE`).
    delta = _TODAY_TX_DELTAS.get(portfolio_id, 0.0)
    return baseline + delta


# Populated by main() before sheet build
_TODAY_TX_DELTAS: dict[str, float] = {}


def fetch_today_tx_delta(sb, portfolio_id: str, initial_capital: float) -> float:
    txns = (
        sb.table("transactions")
        .select("action, weight_in, weight_out, price_out, entry_price_out")
        .eq("portfolio_id", portfolio_id)
        .eq("date", TODAY)
        .execute()
        .data
    )
    delta = 0.0
    for t in txns:
        a = (t.get("action") or "").upper()
        if a == "IN":
            delta -= float(t.get("weight_in") or 0) * initial_capital / 100.0
        elif a in ("TRIM", "OUT"):
            w = float(t.get("weight_out") or 0)
            pru = float(t.get("entry_price_out") or 0)
            p = float(t.get("price_out") or 0)
            if pru > 0:
                delta += (w * initial_capital / 100.0 / pru) * p
        elif a == "SWITCH":
            w_in = float(t.get("weight_in") or 0)
            w_out = float(t.get("weight_out") or 0)
            pru = float(t.get("entry_price_out") or 0)
            p = float(t.get("price_out") or 0)
            if pru > 0:
                delta += (w_out * initial_capital / 100.0 / pru) * p
            delta -= w_in * initial_capital / 100.0
    return delta


# ── Moves sheet ───────────────────────────────────────────────────────────────
def build_moves_sheet(wb: Workbook, portfolio_id: str, pf_row: dict,
                      transactions: list[dict], initial_csv: Path | None):
    sheet_name = f"{portfolio_id.capitalize()} Moves"
    ws = wb.create_sheet(sheet_name)

    # ── Section 1: portfolio metadata ────────────────────────────────────────
    ws["A1"] = f"PORTFOLIO: {portfolio_id.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")

    rows_meta = [
        ("Inception date",     pf_row.get("inception_date")),
        ("Initial capital ($)", pf_row.get("initial_capital")),
        ("Benchmark primary",   f'{pf_row.get("benchmark_primary","")} ({pf_row.get("benchmark_primary_label","")})'),
        ("Benchmark secondary", f'{pf_row.get("benchmark_secondary","")} ({pf_row.get("benchmark_secondary_label","")})'),
        ("Description",         pf_row.get("description")),
    ]
    r = 3
    for label, value in rows_meta:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    # ── Section 2: inception state (from CSV mirror) ─────────────────────────
    r += 1
    ws.cell(row=r, column=1, value="INCEPTION STATE (from snapshots/<pid>/<inception>.csv)").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    headers = ["Ticker", "Weight (%)", "Units", "Entry Price ($)", "Snapshot Type"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    r += 1

    if initial_csv and initial_csv.exists():
        with open(initial_csv) as f:
            reader = csv.DictReader(f)
            for row in reader:
                ws.cell(row=r, column=1, value=row.get("ticker"))
                ws.cell(row=r, column=2, value=float(row.get("weight") or 0))
                ws.cell(row=r, column=3, value=float(row.get("units") or 0))
                ws.cell(row=r, column=4, value=float(row.get("entry_price") or 0))
                ws.cell(row=r, column=5, value=row.get("snapshot_type"))
                for c in range(1, 6):
                    ws.cell(row=r, column=c).border = BORDER
                r += 1
    else:
        ws.cell(row=r, column=1, value="(inception CSV not found)").font = Font(italic=True)
        r += 1

    # ── Section 3: transactions log ──────────────────────────────────────────
    r += 1
    ws.cell(row=r, column=1, value="TRANSACTIONS LOG (chronological — replay these from inception to rebuild current state)").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
    r += 1

    tx_headers = [
        "ID", "Date", "Executed At (UTC)", "Action",
        "Ticker IN", "Weight IN (%)", "Price IN ($)",
        "Ticker OUT", "Weight OUT (%)", "Price OUT ($)", "PRU at exit ($)",
        "Perf %", "Cash delta ($)", "Reason",
    ]
    for c, h in enumerate(tx_headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    r += 1

    ic = float(pf_row.get("initial_capital") or 1_000_000)
    for t in transactions:
        cash_delta = compute_cash_delta(t, ic)
        vals = [
            t.get("id"),
            t.get("date"),
            t.get("executed_at"),
            t.get("action"),
            t.get("ticker_in"),
            float(t["weight_in"])  if t.get("weight_in")  is not None else None,
            float(t["price_in"])   if t.get("price_in")   is not None else None,
            t.get("ticker_out"),
            float(t["weight_out"]) if t.get("weight_out") is not None else None,
            float(t["price_out"])  if t.get("price_out")  is not None else None,
            float(t["entry_price_out"]) if t.get("entry_price_out") is not None else None,
            float(t["perf_pct"])   if t.get("perf_pct")   is not None else None,
            round(cash_delta, 2)   if cash_delta is not None else None,
            t.get("reason"),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
        r += 1

    # Column widths
    widths = {
        "A": 8, "B": 12, "C": 22, "D": 8,
        "E": 10, "F": 12, "G": 12,
        "H": 10, "I": 13, "J": 13, "K": 14,
        "L": 8, "M": 14, "N": 50,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def compute_cash_delta(t: dict, initial_capital: float) -> float | None:
    """Return the $ flow into cash from this transaction (positive = received)."""
    a = (t.get("action") or "").upper()
    if a == "IN":
        w = float(t.get("weight_in") or 0)
        return -w * initial_capital / 100.0
    if a in ("TRIM", "OUT"):
        w = float(t.get("weight_out") or 0)
        pru = float(t.get("entry_price_out") or 0)
        p = float(t.get("price_out") or 0)
        if pru > 0:
            return (w * initial_capital / 100.0 / pru) * p
        return None
    if a == "SWITCH":
        w_in = float(t.get("weight_in") or 0)
        w_out = float(t.get("weight_out") or 0)
        pru = float(t.get("entry_price_out") or 0)
        p = float(t.get("price_out") or 0)
        d = 0.0
        if pru > 0:
            d += (w_out * initial_capital / 100.0 / pru) * p
        d -= w_in * initial_capital / 100.0
        return d
    if a in ("DRIP", "SPLIT"):
        return 0.0
    return None


# ── README sheet ──────────────────────────────────────────────────────────────
def build_readme(wb: Workbook, summaries: list[tuple[str, float, int, int]]):
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "Specula — Daily Holdings Audit Workbook"
    ws["A1"].font = Font(bold=True, size=16)

    ws["A3"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"
    ws["A4"] = f"Source: Supabase tables `daily_holdings` + `transactions` + snapshots/<pid>/<inception>.csv"
    ws["A5"] = f"Live prices for today ({TODAY}): yfinance fast_info.last_price"

    ws["A7"] = "PURPOSE — Audit + Recovery"
    ws["A7"].font = Font(bold=True, size=12)
    ws["A8"] = "If Supabase data is lost or doubted, this file (+ snapshots/ CSVs) is enough to rebuild every portfolio:"
    ws["A9"] = "  1. Read inception state from <Portfolio> Moves → INCEPTION STATE block"
    ws["A10"] = "  2. Replay every row in <Portfolio> Moves → TRANSACTIONS LOG (chronological)"
    ws["A11"] = "  3. Cross-check against <Portfolio> sheet → daily NAV trajectory"

    ws["A13"] = "TABS"
    ws["A13"].font = Font(bold=True, size=12)
    ws["A14"] = "  <Portfolio>           daily NAV + per-ticker (value / shares / price), last row = today live"
    ws["A15"] = "  <Portfolio> Moves     portfolio metadata + inception state + transaction log (full audit)"

    ws["A17"] = "PER-PORTFOLIO SUMMARY"
    ws["A17"].font = Font(bold=True, size=12)
    for i, (name, nav, n_dates, n_txns) in enumerate(summaries, start=18):
        ws.cell(row=i, column=1, value=(
            f"  {name:13} NAV today ~= ${nav:>12,.2f}   |   {n_dates:>3} daily snapshots   |   {n_txns:>2} transactions"
        ))

    ws["A24"] = "REFRESH"
    ws["A24"].font = Font(bold=True, size=12)
    ws["A25"] = "  python build_audit_workbook.py    # re-run any time to regenerate this file with current data"
    ws["A26"] = "  CRITICAL: run after every move + at least once per day post-close US (22h CH)"
    ws["A26"].font = Font(italic=True, color="C00000")

    ws.column_dimensions["A"].width = 120


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    secrets = toml.load(Path(__file__).parent / ".streamlit" / "secrets.toml")
    sb = create_client(secrets["supabase_url"], secrets["supabase_key"])

    wb = Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    summaries: list[tuple[str, float, int, int]] = []
    portfolio_ids = _get_all_portfolio_ids(sb)
    print(f"Portfolios found in DB: {portfolio_ids}")

    for pid in portfolio_ids:
        # Portfolio metadata
        pf = sb.table("portfolios").select("*").eq("id", pid).execute().data
        if not pf:
            print(f"[{pid}] portfolio row not found, skipping")
            continue
        pf_row = pf[0]
        ic = float(pf_row.get("initial_capital") or 1_000_000)

        # daily_holdings
        rows = fetch_holdings(sb, pid)
        if not rows:
            print(f"[{pid}] no daily_holdings, skipping")
            continue

        # Pre-compute today's cash delta (used by holdings sheet)
        _TODAY_TX_DELTAS[pid] = fetch_today_tx_delta(sb, pid, ic)

        # Active positions (for shares lookup on today's row)
        active = (
            sb.table("positions")
            .select("ticker, shares, weight, entry_price")
            .eq("portfolio_id", pid).eq("is_active", True)
            .execute().data
        )

        # Build holdings sheet
        sheet_name, nav_today = build_holdings_sheet(wb, pid, rows, active)
        print(f"[{pid}] holdings sheet '{sheet_name}': "
              f"{len({r['date'] for r in rows})} historical dates + 1 today  "
              f"-> NAV today ~= ${nav_today:,.2f}")

        # Transactions log
        txns = (
            sb.table("transactions").select("*")
            .eq("portfolio_id", pid)
            .order("date").order("id")
            .execute().data
        )
        initial_csv = Path(__file__).parent / "snapshots" / pid / f"{pf_row.get('inception_date')}.csv"
        build_moves_sheet(wb, pid, pf_row, txns, initial_csv)
        print(f"[{pid}] moves sheet built: {len(txns)} transactions, inception CSV "
              f"{'OK' if initial_csv.exists() else 'MISSING'}")

        summaries.append((pid, nav_today, len({r["date"] for r in rows}) + 1, len(txns)))

    build_readme(wb, summaries)
    # Reorder: README first, then per portfolio (holdings, moves)
    desired = ["README"]
    for pid in portfolio_ids:
        for suffix in ["", " Moves"]:
            name = f"{pid.capitalize()}{suffix}"
            if name in wb.sheetnames:
                desired.append(name)
    # openpyxl: reorder via _sheets attribute
    wb._sheets = [wb[name] for name in desired if name in wb.sheetnames]

    wb.save(WORKBOOK_PATH)
    print(f"\nWrote {WORKBOOK_PATH}")

    # Upload to Supabase Storage so a fixed public URL always points to the
    # latest version (no need to git-commit binary file or run local script).
    # Bucket `audit-workbook` must exist and be PUBLIC (configured manually
    # via Supabase Dashboard → Storage → New bucket).
    try:
        with open(WORKBOOK_PATH, "rb") as f:
            file_bytes = f.read()
        # upsert=True overwrites the existing file with the same name
        sb.storage.from_("audit-workbook").upload(
            path="daily_holdings_audit.xlsx",
            file=file_bytes,
            file_options={
                "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "upsert": "true",
            },
        )
        public_url = sb.storage.from_("audit-workbook").get_public_url("daily_holdings_audit.xlsx")
        print(f"Uploaded to Supabase Storage: {public_url}")
    except Exception as e:
        print(f"WARN: Supabase Storage upload failed: {e}")
        print("(workbook still saved locally — manual download possible)")


if __name__ == "__main__":
    main()
