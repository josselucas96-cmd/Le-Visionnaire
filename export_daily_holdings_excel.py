"""Export daily_holdings as an Excel workbook (one sheet per portfolio).

Layout per sheet:
  - Row 1: header
  - Col A: date
  - Col B: NAV total (sum of all tickers' value that day)
  - For each active ticker T: 3 columns (T_value, T_shares, T_price)
  - CASH treated like any ticker (price = 1.0, shares = $ cash)

Output: daily_holdings_audit.xlsx in Streamlit_project/.
"""
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import toml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import create_client


PORTFOLIOS = ["visionnaire", "batisseur", "nakamoto"]


def fetch_portfolio_rows(sb, portfolio_id: str):
    """Return rows sorted by (date, ticker)."""
    return (
        sb.table("daily_holdings")
        .select("date, ticker, shares, price, value")
        .eq("portfolio_id", portfolio_id)
        .order("date")
        .order("ticker")
        .execute()
        .data
    )


def build_sheet(wb: Workbook, portfolio_id: str, rows: list):
    """Pivot rows -> (date × ticker) and write to a new sheet."""
    ws = wb.create_sheet(portfolio_id.capitalize())

    # Group by date / ticker
    by_date: dict[str, dict[str, dict]] = defaultdict(dict)
    tickers_seen: set[str] = set()
    for r in rows:
        by_date[r["date"]][r["ticker"]] = r
        tickers_seen.add(r["ticker"])

    # CASH first, then equity tickers alphabetically
    equity_tickers = sorted(t for t in tickers_seen if t != "CASH")
    ordered_tickers = (["CASH"] if "CASH" in tickers_seen else []) + equity_tickers

    # Header
    header = ["Date", "NAV Total ($)"]
    for tk in ordered_tickers:
        header.extend([f"{tk}_value", f"{tk}_shares", f"{tk}_price"])
    ws.append(header)

    bold = Font(bold=True)
    grey_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = grey_fill
        cell.alignment = Alignment(horizontal="center")

    # Body rows
    for d in sorted(by_date.keys()):
        row = [d]
        nav_total = sum(float(r["value"] or 0) for r in by_date[d].values())
        row.append(round(nav_total, 2))
        for tk in ordered_tickers:
            r = by_date[d].get(tk)
            if r is None:
                row.extend([None, None, None])
            else:
                row.extend([
                    round(float(r["value"] or 0), 2),
                    round(float(r["shares"] or 0), 8),
                    round(float(r["price"] or 0), 4),
                ])
        ws.append(row)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    for i, tk in enumerate(ordered_tickers):
        base_col = 3 + i * 3
        ws.column_dimensions[get_column_letter(base_col)].width = 14      # value
        ws.column_dimensions[get_column_letter(base_col + 1)].width = 14  # shares
        ws.column_dimensions[get_column_letter(base_col + 2)].width = 11  # price

    # Freeze header + date column
    ws.freeze_panes = "C2"

    return ws


def main():
    secrets = toml.load(Path(__file__).parent / ".streamlit" / "secrets.toml")
    sb = create_client(secrets["supabase_url"], secrets["supabase_key"])

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    summary_lines = []
    for pid in PORTFOLIOS:
        rows = fetch_portfolio_rows(sb, pid)
        if not rows:
            print(f"[{pid}] no rows, skipping")
            continue
        build_sheet(wb, pid, rows)
        dates = sorted({r["date"] for r in rows})
        tickers = sorted({r["ticker"] for r in rows})
        nav_last = sum(float(r["value"] or 0) for r in rows if r["date"] == dates[-1])
        summary_lines.append(
            f"  {pid:13} {len(rows):4} rows  "
            f"{len(dates):3} dates ({dates[0]} -> {dates[-1]})  "
            f"{len(tickers):3} tickers  NAV@{dates[-1]}=${nav_last:,.2f}"
        )
        print(summary_lines[-1])

    # Index sheet at the front
    info = wb.create_sheet("README", 0)
    info["A1"] = "daily_holdings audit"
    info["A1"].font = Font(bold=True, size=14)
    info["A3"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"
    info["A4"] = "Source: Supabase table daily_holdings (real fund accounting)"
    info["A6"] = "Columns per ticker:"
    info["A7"] = "  T_value  = shares × price (= $ value on that date)"
    info["A8"] = "  T_shares = number of shares held"
    info["A9"] = "  T_price  = close price used"
    info["A11"] = "CASH is treated as a position with price=1.0 and shares=$ amount"
    info["A12"] = "NAV Total ($) = sum of all _value columns for that date"
    info["A14"] = "Per portfolio summary:"
    for i, line in enumerate(summary_lines, start=15):
        info[f"A{i}"] = line
    info.column_dimensions["A"].width = 110

    out = Path(__file__).parent / "daily_holdings_audit.xlsx"
    wb.save(out)
    print(f"\nWrote {out}")


if __name__ == "__main__":
    main()
