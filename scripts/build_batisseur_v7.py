"""Generate LeBatisseur_Portfolio_v7.xlsx from the agreed v7 composition.

26 positions + 5% cash · UCITS V compliant · inception May 2, 2026.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"c:\Users\USER\Desktop\Projet Claude\Claude Racine\01_THE PORTFOLIO PROJECT\Les portefeuilles\Le Bâtisseur\LeBatisseur_Portfolio_v7.xlsx")

# ── Composition ───────────────────────────────────────────────────────────────
# (ticker, company, country, sector_own, gics, strate, weight, cagr, rating, thesis)
POSITIONS = [
    # OBVIOUS — top conviction (8 names, 46%)
    ("NVDA",   "NVIDIA",                  "US",     "Semiconductors",      "Information Technology",  "Obvious",          0.08, 0.532, "CHEAP", "AI infrastructure leader, FwPE 24x. +1pp vs S&P NVDA weight (~7%) — contained."),
    ("CSU.TO", "Constellation Software",  "Canada", "Software",            "Information Technology",  "Obvious",          0.07, 0.30,  "DEAL",  "Serial acquirer VMS, Mark Leonard. ~30% CAGR 15Y, FwPE 16x DEAL pure."),
    ("AMZN",   "Amazon",                  "US",     "Internet/Cloud",      "Consumer Discretionary",  "Obvious",          0.06, 0.249, "CHEAP", "AWS + retail + ads compounder. Best EV/GM/GR ratio."),
    ("META",   "Meta Platforms",          "US",     "Internet/Media",      "Communication Services",  "Obvious",          0.06, 0.203, "CHEAP", "AI capex bet — if right, asymmetric upside; if wrong, FCF explodes (cashflow yield)."),
    ("MELI",   "MercadoLibre",            "LatAm",  "Internet/Fintech",    "Consumer Discretionary",  "Obvious",          0.05, 0.274, "DEAL",  "LATAM ecosystem + demographic tailwind. Most consumer-cyclical of the 4 platforms."),
    ("LLY",    "Eli Lilly",               "US",     "Pharma",              "Healthcare",              "Obvious",          0.05, 0.231, "FAIR",  "GLP-1 leader (Mounjaro/Zepbound) + neuro/oncology pipeline. Excellent risk/reward."),
    ("JPM",    "JPMorgan Chase",          "US",     "Banking",             "Financials",              "Obvious",          0.045, 0.168, "FAIR",  "Tom Lee thesis: bank → tech via AI + crypto = cost compression + tech multiple."),
    ("BABA",   "Alibaba",                 "China",  "Internet/Cloud",      "Consumer Discretionary",  "Obvious",          0.04, None,  "CHEAP", "China contrarian, FwPE 11x. AI capex (Qwen) underappreciated, buybacks accelerating."),
    # HAUTE QUALITÉ — high quality with caveats (9 names, 28%)
    ("MSFT",   "Microsoft",               "US",     "Software",            "Information Technology",  "Haute Qualité",    0.04, 0.217, "CHEAP", "Office/Azure backbone. Only OpenAI deal structure as overhang."),
    ("V",      "Visa",                    "US",     "Payment Networks",    "Financials",              "Haute Qualité",    0.04, 0.19,  "FAIR",  "Network effect monopoly. Stablecoin question hedged by CRCL position."),
    ("BSX",    "Boston Scientific",       "US",     "Med Devices",         "Healthcare",              "Haute Qualité",    0.04, 0.149, "CHEAP", "Cardio/EP leader. Switching cost moat (training, hospital adoption stickiness)."),
    ("FICO",   "Fair Isaac",              "US",     "Data/Analytics",      "Information Technology",  "Haute Qualité",    0.03, 0.27,  "DEAL",  "Credit scoring monopoly. Post -41% drawdown while fundamentals accelerate."),
    ("VRSK",   "Verisk Analytics",        "US",     "Data/Analytics",      "Industrials",             "Haute Qualité",    0.03, 0.15,  "FAIR",  "Insurance data monopoly. Recurring SaaS-like revenue."),
    ("ISRG",   "Intuitive Surgical",      "US",     "Med Devices",         "Healthcare",              "Haute Qualité",    0.03, 0.257, "CULT",  "Robotic surgery monopoly. CULT mérité — recurring revenue model."),
    ("ULTA",   "Ulta Beauty",             "US",     "Specialty Retail",    "Consumer Discretionary",  "Haute Qualité",    0.03, 0.10,  "CHEAP", "ROIC 24%, FwPE 18x, 25% below historical median. Hybrid prestige+mass moat."),
    ("RACE",   "Ferrari",                 "Italy",  "Luxury Auto",         "Consumer Discretionary",  "Haute Qualité",    0.02, 0.12,  "DEAL",  "CULT brand at FwPE 30x (32% below 5Y avg). Rev growth decelerating to 4.66% — sized small."),
    ("RMS.PA", "Hermès",                  "France", "Luxury",              "Consumer Discretionary",  "Haute Qualité",    0.02, 0.21,  "CULT",  "Pricing power culte 5-7%/yr. Rev growth 1.92% TTM (China weakness) — sized small until reaccel."),
    # DIVERSIFICATION — quality for portfolio balance (7 names, 17%)
    ("EL.PA",  "EssilorLuxottica",        "France", "Med Devices/Eyewear", "Consumer Discretionary",  "Diversification",  0.03, 0.10,  "FAIR",  "Ray-Ban Meta smart glasses catalyst. Fairly valued — held for diversification."),
    ("WM",     "Waste Management",        "US",     "Industrials/Waste",   "Industrials",             "Diversification",  0.03, 0.10,  "FAIR",  "Landfills regulatory monopoly. Physical moat. ROIC 9-11% accepted exception."),
    ("NVO",    "Novo Nordisk",            "Denmark","Pharma",              "Healthcare",              "Diversification",  0.03, 0.12,  "DEAL*", "GLP-1 recovery play. FwPE 13x. DEAL conditional on Wegovy stabilization + pipeline."),
    ("VICI",   "VICI Properties",         "US",     "REIT/Casino",         "Real Estate",             "Tactical",         0.02, 0.08,  "CHEAP", "REIT casino properties, 6% div, P/AFFO 13x. Fills Real Estate gap (was 0%). Tactical (sector rotation play)."),
    ("IBE.MC", "Iberdrola",               "Spain",  "Regulated Utility/Renewables", "Utilities",      "Diversification",  0.02, 0.10,  "FAIR",  "Spanish multinational utility, ~40% renewables. AI data center power demand catalyst. Stock +30% YTD — sized 2% cautiously, monitor for re-rating risk if AI narrative shifts. Fills Utilities gap (was 0%)."),
    ("DPZ",    "Domino's Pizza",          "US",     "Consumer Disc",       "Consumer Discretionary",  "Diversification",  0.02, 0.10,  "CHEAP", "Multiple expansion catalyst: FwPE 16x → 20-25x lorsque sentiment consommateur s'améliore. Quality franchise + ROIC fabuleux + Buffett conviction comme back-stop."),
    ("STZ",    "Constellation Brands",    "US",     "Beverages",           "Consumer Staples",        "Diversification",  0.02, None,  "CHEAP", "Peur excessive sur tarifs → adaptation entreprise via brand strength Modelo/Corona + pricing power. Turmoil court terme, repricing attendu une fois adaptation digérée. FwPE 13x = lowest in book."),
    # TACTICAL (2 names, 4%)
    ("ZM",     "Zoom",                    "US",     "Software",            "Information Technology",  "Tactical",         0.02, None,  "DEAL",  "Anthropic SoTP — ~30% of EV is the Anthropic stake at $900B valuation."),
    ("CRCL",   "Circle",                  "US",     "Crypto Infrastructure","Financials",             "Tactical",         0.02, None,  "FAIR",  "USDC stablecoin issuer. Tactical: regulation pending. Compo Visa+Circle = stablecoin hedge."),
    # CASH
    ("Cash",   "Cash buffer",             "—",      "Cash",                "Cash",                    "Cash",             0.055, None, "—",     "Opportunistic optionality (per IPS spirit)."),
]

CHANGES_VS_V6 = [
    ("Drop", "TSM",   "0.03", "User dislikes TSMC; same theme as NVDA (foundry → GPU), not a real hedge."),
    ("Drop", "LNG",   "0.02", "User: 'pas de conviction, conseil de Claude'. Violates 'no genuine edge' rule."),
    ("Drop", "NFLX",  "0.03", "FwPE 38x for 27% CAGR decelerating. Risk/reward dégradé."),
    ("Add",  "VICI",  "0.02", "REIT casino. Fills Real Estate 0% gap (PM concern). Already on watchlist."),
    ("Add",  "IBE.MC","0.02", "Iberdrola — Spanish utility + 40% renewables. Fills Utilities 0% gap. No 'AI proxy' bias. Sized 2% cautiously (+30% YTD)."),
    ("↓",    "MELI",  "0.06→0.05", "Conviction forte mais valuation déjà chère."),
    ("↑",    "LLY",   "0.03→0.05", "Excellent risk/reward GLP-1 + pipeline."),
    ("↑",    "JPM",   "0.04→0.045", "Tom Lee tech-thesis. Capped at 4.5% to keep sum >5% comfortably below alert."),
    ("↑",    "WM",    "0.02→0.03", "Moat physique unique mérite plus que 2%."),
    ("↑",    "CRCL",  "0.02→0.02 (Tactical)", "Reclassed Tactical (was Élite). Conviction présente mais binaire/régulation."),
    ("↓",    "ZM",    "0.03→0.02", "User: 'on peut mettre 2% si tu veux'."),
    ("Cash", "Cash",  "0.02→0.05", "Reconstruction du buffer per IPS spirit."),
]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# Styles
TITLE_FONT = Font(bold=True, size=14, color="1F2937")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0E1117")
SUBHEADER_FONT = Font(bold=True, size=11, color="92400E")
SUBHEADER_FILL = PatternFill("solid", fgColor="FEF3C7")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def write_header(ws, row, headers, fill=HEADER_FILL, font=HEADER_FONT):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER


def auto_width(ws, max_w=60):
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 2, max_w)


# ─── Sheet 1: Notes ───────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Notes"
ws["A1"] = "Le Bâtisseur — Portfolio Notes (AI Handoff)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Specula house · Quality Compounders + Tactical · UCITS V compliant · 26 positions · v7 (refined from v6)"

NOTES_SECTIONS = [
    ("  1. CONTEXT — Specula House", [
        ("Owner", "Lucas, builder of paper portfolios under the 'Specula' house brand."),
        ("Three sub-portfolios", "(1) Le Visionnaire — High Conviction Equity, Nasdaq 100 benchmark, no UCITS, violet identity. (2) Le Nakamoto — Digital Asset Treasuries (BTC), BTC benchmark, Very High risk, orange identity. (3) Le Bâtisseur — Quality Compounders + Tactical, S&P 500 benchmark, UCITS V compliant, gold identity."),
        ("Mission", "The 'respectability portfolio' — must beat S&P 500 over 5+ years through quality compounders, with strict risk discipline. Tactical layer adds asymmetric special situations."),
        ("Inception", "May 2, 2026 (per IPS Finale)."),
    ]),
    ("  2. WHAT CHANGED v6 → v7", [
        (f"{c[0]} {c[1]}", f"{c[2]} — {c[3]}") for c in CHANGES_VS_V6
    ]),
    ("  3. CONVICTION TIERS (user's framework)", [
        ("Obvious (45.5%)", "8 noms top conviction. NVDA 8, CSU 7, AMZN 6, META 6, MELI 5, LLY 5, JPM 4.5, BABA 4."),
        ("Haute Qualité (28%)", "9 noms haute qualité avec quelques caveats valuation/cycle. MSFT, V, BSX, FICO, VRSK, ISRG, ULTA, RACE, RMS."),
        ("Diversification (17%)", "7 noms pour combler les trous sectoriels et équilibrer le profil. EL, WM, NVO, VICI, IBE.MC, DPZ, STZ."),
        ("Tactical (4%)", "2 noms à thèse à catalyseur défini. ZM (Anthropic SoTP), CRCL (stablecoin regulation)."),
        ("Cash (5.5%)", "Buffer optionalité reconstruit (vs 2% v6)."),
    ]),
    ("  4. RATING SYSTEM", [
        ("DEAL",  "Significantly mispriced — quality at unusually low multiple. Ex: CSU FwPE 16x for 30% CAGR, FICO FwPE 24x for 27% growth."),
        ("CHEAP", "Below fair value given growth/quality. Ex: NVDA, META, AMZN, BABA."),
        ("FAIR",  "In line with quality/growth — pay full price for full quality. Ex: V, JPM, LLY, NG.L."),
        ("CULT",  "Paying for scarcity/quality (justified) — multi-decade compounders never cheap. Ex: ISRG, RMS."),
        ("DEAL*", "DEAL conditional on specific scenario. Ex: NVO (GLP-1 recovery)."),
    ]),
    ("  5. UCITS V COMPLIANCE (live)", [
        ("Rule 1 — Single issuer cap", "No single position >10% of NAV. Max v7: NVDA 8% (well below). Personal trim alert at 9.5%."),
        ("Rule 2 — Sum positions >5%", "Must not exceed 40%. Sum strictly >5% in v7: NVDA 8 + CSU 7 + AMZN 6 + META 6 = 27%. Comfortable headroom."),
        ("Rule 2 — Note", "MELI/LLY/JPM at exactly 5% don't count toward the 40% sum (UCITS rule is strictly >5%). They're at the threshold — first growth tick puts them over."),
        ("Rule 3 — Min diversification", "At least 16 positions. v7: 26 — comfortable."),
    ]),
    ("  6. ADDRESSED PM CONCERNS vs v6", [
        ("Real Estate 0%", "Solved via VICI 2% (casino REIT, 6% div, regulatory moat)."),
        ("Utilities 0%", "Solved via IBE.MC 2% (Iberdrola, Spanish utility + 40% renewables). Avoids 'AI utility proxy' (CEG-type) bias. Note: stock +30% YTD — sized cautiously."),
        ("Tactical 21% (v6)", "Reduced to 4% (ZM + CRCL). Tighter monitoring, IPS-cleaner."),
        ("Cash 2% (v6)", "Reconstituted to 5% per IPS spirit ('Better to hold cash than fill mediocre tactical bets')."),
        ("NFLX kept (v6)", "Dropped — 38x FwPE for decelerating 27% CAGR, risk/reward dégradé."),
    ]),
    ("  7. RESIDUAL TENSIONS (knowingly accepted)", [
        ("Energy 0%, Materials 0%", "Defensible — no quality compounder edge in commodities. IPS-aligned ('refuse positions without genuine edge')."),
        ("Consumer Discretionary 27% (vs S&P 10%)", "+17pp overweight, biggest sector bet. Driven by AMZN+MELI+BABA+ULTA+EL+RACE+RMS+DPZ. Cyclical risk concentration."),
        ("RACE/RMS rev growth decelerating", "RACE 4.66% TTM, RMS 1.92% TTM. Sized small (2% each). No upgrade until reacceleration signal (China + China + new product cycle)."),
        ("AI capex sensitive ~24%", "NVDA 8 + AMZN 6 + META 6 + MSFT 4. If AI capex narrative breaks, hyperscaler FCF explodes (internal hedge)."),
        ("Healthcare cluster 15%", "LLY+BSX+ISRG+NVO. Common reimbursement/regulatory exposure but each has distinct moat."),
    ]),
    ("  8. POSITION RATIONALES — see Portfolio sheet", [
        ("Detail", "Each position has thesis, rating, country, sector, GICS in the Portfolio sheet."),
    ]),
    ("  9. NEXT STEPS", [
        ("Stock papers priority", "(1) CSU.TO DEAL serial acquirer. (2) MELI EM thesis. (3) FICO DEAL exceptional. (4) JPM Tom Lee tech-thesis. (5) NVO GLP-1 recovery."),
        ("Watchlist re-entries", "Sartorius (DIM.PA), Stryker (SYK), Cadence/Synopsys, Salesforce, Lindt, Hexagon RemainCo, Freeport, Cameco, MarketAxess, Kingspan, United Rentals."),
        ("Quarterly rebalance discipline", "(a) Single ≥9.5% → trim to 7-8%. (b) Sum >5% reaches 39% → trim winners. (c) RACE/RMS upgrade if rev growth reaccelerates. (d) STZ thesis check (sales -10% organic FY26 = warning)."),
        ("Site activation", "Insert positions in Supabase with portfolio_id='batisseur', inception 2026-05-02."),
    ]),
]

row = 4
for section_title, rows in NOTES_SECTIONS:
    c = ws.cell(row=row, column=1, value=section_title)
    c.font = SUBHEADER_FONT
    c.fill = SUBHEADER_FILL
    row += 1
    for k, v in rows:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1  # spacer

ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 110

# ─── Sheet 2: Portfolio ──────────────────────────────────────────────────────
ws = wb.create_sheet("Portfolio")
ws["A1"] = "Portfolio Composition — v7"
ws["A1"].font = TITLE_FONT
ws["A2"] = "26 positions + cash · grouped by Strate · UCITS V compliant"

headers = ["Ticker", "Company", "Country", "Sector", "GICS", "Strate", "Weight", "CAGR", "Rating", "Thesis"]
write_header(ws, 4, headers)

# Group by Strate
TIER_HEADERS = {
    "Obvious":         "OBVIOUS — Top conviction (45.5%)",
    "Haute Qualité":   "HAUTE QUALITÉ — High quality with caveats (28%)",
    "Diversification": "DIVERSIFICATION — Sector balance + IPS coverage (17%)",
    "Tactical":        "TACTICAL — Catalyst-driven (4%)",
    "Cash":            "CASH (5.5%)",
}
TIER_ORDER = ["Obvious", "Haute Qualité", "Diversification", "Tactical", "Cash"]

row = 5
for tier in TIER_ORDER:
    c = ws.cell(row=row, column=1, value=TIER_HEADERS[tier])
    c.font = SUBHEADER_FONT
    c.fill = SUBHEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1
    for p in [pp for pp in POSITIONS if pp[5] == tier]:
        ticker, company, country, sector, gics, strate, weight, cagr, rating, thesis = p
        ws.cell(row=row, column=1, value=ticker)
        ws.cell(row=row, column=2, value=company)
        ws.cell(row=row, column=3, value=country)
        ws.cell(row=row, column=4, value=sector)
        ws.cell(row=row, column=5, value=gics)
        ws.cell(row=row, column=6, value=strate)
        wc = ws.cell(row=row, column=7, value=weight)
        wc.number_format = "0.00%"
        ws.cell(row=row, column=8, value=cagr if cagr else "n/a")
        if cagr:
            ws.cell(row=row, column=8).number_format = "0.0%"
        ws.cell(row=row, column=9, value=rating)
        ws.cell(row=row, column=10, value=thesis).alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

# TOTAL row
ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=row, column=2, value="26 positions + cash buffer").font = Font(bold=True)
total = ws.cell(row=row, column=7, value=sum(p[6] for p in POSITIONS))
total.number_format = "0.00%"
total.font = Font(bold=True)

ws.column_dimensions["A"].width = 9
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 9
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 26
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 9
ws.column_dimensions["H"].width = 8
ws.column_dimensions["I"].width = 9
ws.column_dimensions["J"].width = 70

# ─── Sheet 3: Breakdowns ────────────────────────────────────────────────────
ws = wb.create_sheet("Breakdowns")
ws["A1"] = "Breakdowns — v7"
ws["A1"].font = TITLE_FONT

# GICS
gics_agg = {}
for p in POSITIONS:
    gics_agg.setdefault(p[4], []).append((p[0], p[6]))

ws.cell(row=3, column=1, value="GICS Sectors").font = SUBHEADER_FONT
ws.cell(row=3, column=1).fill = SUBHEADER_FILL
write_header(ws, 4, ["GICS Sector", "Weight", "Positions"])
gics_order = ["Consumer Discretionary", "Information Technology", "Healthcare", "Financials",
              "Communication Services", "Industrials", "Consumer Staples", "Real Estate",
              "Utilities", "Cash"]
row = 5
for sector in gics_order:
    if sector in gics_agg:
        positions = gics_agg[sector]
        total_w = sum(w for _, w in positions)
        ws.cell(row=row, column=1, value=sector)
        wc = ws.cell(row=row, column=2, value=total_w)
        wc.number_format = "0.0%"
        ws.cell(row=row, column=3, value=", ".join(f"{t} {int(w*100)}" for t, w in positions))
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

# Economic Exposure (own framework)
row += 2
ws.cell(row=row, column=1, value="Economic Exposure (own framework)").font = SUBHEADER_FONT
ws.cell(row=row, column=1).fill = SUBHEADER_FILL
row += 1
write_header(ws, row, ["Bucket", "Weight", "Positions"])
row += 1

ECON_BUCKETS = [
    ("Enterprise Software",                ["CSU.TO", "MSFT", "FICO", "ZM"]),
    ("Financial Infrastructure",            ["JPM", "V", "VRSK"]),
    ("Digital Platforms (multi-engine)",    ["AMZN", "META"]),
    ("Med Devices",                          ["BSX", "ISRG", "EL.PA"]),
    ("Consumer Internet (cyclical EM)",     ["MELI", "BABA"]),
    ("Semiconductors / AI Infra",           ["NVDA"]),
    ("Pharma & GLP-1",                       ["LLY", "NVO"]),
    ("Consumer Specialty & Brands",         ["ULTA", "DPZ", "STZ"]),
    ("Cash",                                 ["Cash"]),
    ("Luxury & Premium",                     ["RACE", "RMS.PA"]),
    ("Industrials Defensive",                ["WM"]),
    ("Real Estate",                          ["VICI"]),
    ("Utilities",                            ["NG.L"]),
    ("Crypto Infrastructure",                ["CRCL"]),
]
pos_by_t = {p[0]: p[6] for p in POSITIONS}
for bucket, tickers in ECON_BUCKETS:
    weights = [(t, pos_by_t.get(t, 0)) for t in tickers]
    total_w = sum(w for _, w in weights)
    ws.cell(row=row, column=1, value=bucket)
    wc = ws.cell(row=row, column=2, value=total_w)
    wc.number_format = "0.0%"
    ws.cell(row=row, column=3, value=", ".join(f"{t} {int(w*100)}" for t, w in weights))
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

# Geographic
row += 2
ws.cell(row=row, column=1, value="Geographic").font = SUBHEADER_FONT
ws.cell(row=row, column=1).fill = SUBHEADER_FILL
row += 1
write_header(ws, row, ["Region", "Weight", "Positions"])
row += 1
geo_agg = {}
for p in POSITIONS:
    geo_agg.setdefault(p[2], []).append((p[0], p[6]))
geo_order = ["US", "Canada", "LatAm", "China", "France", "Italy", "Denmark", "UK", "—"]
for geo in geo_order:
    if geo in geo_agg:
        positions = geo_agg[geo]
        total_w = sum(w for _, w in positions)
        ws.cell(row=row, column=1, value=geo if geo != "—" else "Cash")
        wc = ws.cell(row=row, column=2, value=total_w)
        wc.number_format = "0.0%"
        ws.cell(row=row, column=3, value=", ".join(f"{t} {int(w*100)}" for t, w in positions))
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

# Rating
row += 2
ws.cell(row=row, column=1, value="Rating Distribution").font = SUBHEADER_FONT
ws.cell(row=row, column=1).fill = SUBHEADER_FILL
row += 1
write_header(ws, row, ["Rating", "Weight", "Positions"])
row += 1
rat_agg = {}
for p in POSITIONS:
    rat_agg.setdefault(p[8], []).append((p[0], p[6]))
rat_order = ["DEAL", "DEAL*", "CHEAP", "FAIR", "CULT", "—"]
for rating in rat_order:
    if rating in rat_agg:
        positions = rat_agg[rating]
        total_w = sum(w for _, w in positions)
        ws.cell(row=row, column=1, value=rating)
        wc = ws.cell(row=row, column=2, value=total_w)
        wc.number_format = "0.0%"
        ws.cell(row=row, column=3, value=", ".join(f"{t} {int(w*100)}" for t, w in positions))
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

# Strate
row += 2
ws.cell(row=row, column=1, value="Strate (conviction tiers)").font = SUBHEADER_FONT
ws.cell(row=row, column=1).fill = SUBHEADER_FILL
row += 1
write_header(ws, row, ["Strate", "Weight", "# Positions"])
row += 1
str_agg = {}
for p in POSITIONS:
    str_agg.setdefault(p[5], []).append((p[0], p[6]))
for strate in TIER_ORDER:
    if strate in str_agg:
        positions = str_agg[strate]
        total_w = sum(w for _, w in positions)
        ws.cell(row=row, column=1, value=strate)
        wc = ws.cell(row=row, column=2, value=total_w)
        wc.number_format = "0.0%"
        ws.cell(row=row, column=3, value=len(positions))
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 11
ws.column_dimensions["C"].width = 80

# ─── Sheet 4: UCITS ──────────────────────────────────────────────────────────
ws = wb.create_sheet("UCITS")
ws["A1"] = "UCITS V Compliance — Live Check"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Règle 5/10/40 · v7 composition"

write_header(ws, 4, ["Rule", "Current Value", "Limit", "Headroom", "Status"])

# Compute
all_pos = [p for p in POSITIONS if p[5] != "Cash"]
max_w = max(p[6] for p in all_pos)
sum_above_5 = sum(p[6] for p in all_pos if p[6] > 0.05)  # strictly > 5%
n_above_5 = sum(1 for p in all_pos if p[6] > 0.05)
n_total = len(all_pos)

ucits_rules = [
    ("Cap individuel max",            max_w, 0.10, 0.10 - max_w, "OK" if max_w <= 0.10 else "BREACH"),
    ("Somme positions >5%",            sum_above_5, 0.40, 0.40 - sum_above_5, "OK" if sum_above_5 <= 0.40 else "BREACH"),
    ("Nombre positions >5%",           n_above_5, "—", "—", "INFO"),
    ("Total positions",                n_total, "min 16", "—", "OK" if n_total >= 16 else "BREACH"),
]
row = 5
for rule, cur, lim, hr, status in ucits_rules:
    ws.cell(row=row, column=1, value=rule)
    cell_cur = ws.cell(row=row, column=2, value=cur)
    if isinstance(cur, float):
        cell_cur.number_format = "0.0%"
    cell_lim = ws.cell(row=row, column=3, value=lim)
    if isinstance(lim, float):
        cell_lim.number_format = "0.0%"
    cell_hr = ws.cell(row=row, column=4, value=hr)
    if isinstance(hr, float):
        cell_hr.number_format = "0.0%"
    cell_st = ws.cell(row=row, column=5, value=status)
    if status == "OK":
        cell_st.fill = PatternFill("solid", fgColor="D1FAE5")
    elif status == "BREACH":
        cell_st.fill = PatternFill("solid", fgColor="FEE2E2")
    for c in range(1, 6):
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

# Personal trim discipline
row += 2
ws.cell(row=row, column=1, value="Personal Trim Discipline (stricter than UCITS)").font = SUBHEADER_FONT
ws.cell(row=row, column=1).fill = SUBHEADER_FILL
row += 1
write_header(ws, row, ["Threshold", "Action", "Trigger", "Personal target", "Why"])
row += 1
discipline = [
    ("Position ≤ 8%",          "Hold",            "—",                "—",            "Within target zone"),
    ("Position 8–9%",          "Watch",           "No add",           "—",            "Approaching cap"),
    ("Position 9–9.5%",        "Trim alert",      "Prepare sell",     "Target 7–8%",  "Pre-emptive"),
    ("Position ≥ 9.5%",        "Trim immediate",  "Sell to 7–8%",     "Hard limit",   "Protect against gap up"),
    ("Position ≥ 10%",         "BREACH",          "Sell within 30d",  "Mandatory",    "UCITS hard rule"),
    ("Sum >5% ≥ 37%",          "Trim alert",      "Trim winners",     "Target 35%",   "Pre-emptive"),
    ("Sum >5% ≥ 39%",          "Trim immediate",  "Trim winners",     "Target 35%",   "Approaching limit"),
    ("Sum >5% ≥ 40%",          "BREACH",          "Trim within 30d",  "Mandatory",    "UCITS hard rule"),
]
for d in discipline:
    for col, val in enumerate(d, 1):
        ws.cell(row=row, column=col, value=val).border = THIN_BORDER
    row += 1

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 17
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 28

# ─── Sheet 5: Watchlist ──────────────────────────────────────────────────────
ws = wb.create_sheet("Watchlist")
ws["A1"] = "Watchlist & Stocks Studied"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Active monitoring + permanently studied & passed"

ws.cell(row=4, column=1, value="Active Watchlist — entry conditions defined").font = SUBHEADER_FONT
ws.cell(row=4, column=1).fill = SUBHEADER_FILL
write_header(ws, 5, ["Ticker", "Company", "Country", "Sector", "Entry Trigger", "Target Position", "Rationale"])

watchlist_active = [
    ("LISP.SW", "Lindt & Sprüngli",   "Switzerland", "Consumer/Luxury",  "Correction -15-20%",       "2-3% Qualité",    "CULT mérité mais MS bear case. Wait better entry."),
    ("SYK",     "Stryker",            "US",          "Med Devices",       "Better entry",             "2-3% Qualité",    "Quality compounder med devices. Watch for overlap with BSX/ISRG."),
    ("CDNS",    "Cadence Design",     "US",          "EDA Software",      "Correction tech",          "2-3% Qualité",    "Duopole EDA software."),
    ("SNPS",    "Synopsys",           "US",          "EDA Software",      "Post Ansys digestion",     "2-3% Qualité",    "Alternative à Cadence."),
    ("CRM",     "Salesforce",         "US",          "Software/CRM",      "Si concentration tech baisse", "3-4% Solide", "FwPE 14x CHEAP, Agentforce catalyst."),
    ("HEXA-B",  "Hexagon RemainCo",   "Sweden",      "Industrial Tech",   "Post spin-off Q3 2026",    "2% Tactical",     "RemainCo simplifié, special situation."),
    ("FCX",     "Freeport-McMoRan",   "US",          "Copper Mining",     "Correction -25%",          "2% Tactical",     "Cuivre least extended grands miniers."),
    ("CCJ",     "Cameco",             "Canada",      "Uranium",           "Correction significative", "2% Tactical",     "Quality uranium producer."),
    ("DIM.PA",  "Sartorius Stedim",   "France",      "Bioprocessing",     "Goldman + earnings",       "3% Qualité",      "Re-entry timing better. Recovery thesis."),
    ("URI",     "United Rentals",     "US",          "Industrials/Rental","Correction -15-20%",       "2-3% Tactical",   "AI data center capex play."),
    ("MKTX",    "MarketAxess",        "US",          "Bond Trading",      "FwPE <20 + share recovery","3% Solide",       "Bond trading platform."),
    ("KGP",     "Kingspan Group",     "Ireland",     "Building Products", "Better timing",            "2-3% Solide",     "Insulation leader, family-led serial acquirer."),
]
row = 6
for w in watchlist_active:
    for col, val in enumerate(w, 1):
        ws.cell(row=row, column=col, value=val).border = THIN_BORDER
    row += 1

# Studied & Passed (concise — link to full v6 list)
row += 2
ws.cell(row=row, column=1, value="Studied & Passed — NOT to revisit unless thesis changes").font = SUBHEADER_FONT
ws.cell(row=row, column=1).fill = SUBHEADER_FILL
row += 1
write_header(ws, row, ["Ticker", "Company", "Country", "Sector", "Status", "Why"])
row += 1
passed = [
    ("AAPL",    "Apple",          "US",   "Tech Hardware", "Passed",        "FwPE 32x for 5% growth. Valuation refused."),
    ("GOOGL",   "Alphabet",       "US",   "Internet",      "Passed",        "FwPE 31x. AI search risk + valuation."),
    ("ORCL",    "Oracle",         "US",   "Software",      "Passed",        "FwPE 30x unproven cloud transition."),
    ("AVGO",    "Broadcom",       "US",   "Semis",         "Passed",        "TSM/NVDA preferred."),
    ("TSLA",    "Tesla",          "US",   "Auto/AI",       "Passed",        "Le Visionnaire candidate, not Bâtisseur."),
    ("BRK.B",   "Berkshire",      "US",   "Financials",    "Passed",        "No growth engine."),
    ("CEG",     "Constellation Energy", "US", "Utilities", "Passed",        "AI proxy déguisé en utility. NG.L preferred."),
    ("BARN",    "Barry Callebaut","Switzerland", "Cocoa",  "Passed",        "Profit warning April 2026. Thesis broken."),
    ("CVX",     "Chevron",        "US",   "Energy",        "Passed",        "Low conviction. No edge in oil."),
    ("NOW",     "ServiceNow",     "US",   "Software",      "Passed",        "Cut to reduce software cluster."),
    ("TTD",     "The Trade Desk", "US",   "AdTech",        "Passed",        "Visionnaire candidate."),
    ("INTU",    "Intuit",         "US",   "Software",      "Passed",        "CULT trop cher (FwPE 35x+)."),
    ("KEYN",    "Keyence",        "Japan","Industrial Sensors", "Passed",   "CULT permanently 45x+."),
    ("MCD",     "McDonald's",     "US",   "QSR",           "Passed",        "Sub-S&P CAGR."),
    ("PANW",    "Palo Alto",      "US",   "Cybersecurity", "Passed",        "FwPE 50x trop cher."),
    ("MSTR",    "MicroStrategy",  "US",   "BTC Treasury",  "Passed",        "Nakamoto overlap."),
    # v7 EXITS
    ("TSM",     "Taiwan Semi",    "Taiwan", "Foundry",     "EXITED v6→v7",  "Same theme as NVDA, not a hedge. User dislikes."),
    ("LNG",     "Cheniere Energy","US",   "LNG",           "EXITED v6→v7",  "No genuine edge admitted. IPS-violating."),
    ("NFLX",    "Netflix",        "US",   "Streaming",     "EXITED v6→v7",  "FwPE 38x for decelerating 27% CAGR."),
]
for p in passed:
    for col, val in enumerate(p, 1):
        ws.cell(row=row, column=col, value=val).border = THIN_BORDER
    row += 1

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 32
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 60

# ── Save ─────────────────────────────────────────────────────────────────────
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Total invested: {sum(p[6] for p in POSITIONS if p[5] != 'Cash'):.0%}")
print(f"Cash: {sum(p[6] for p in POSITIONS if p[5] == 'Cash'):.0%}")
print(f"Positions: {len([p for p in POSITIONS if p[5] != 'Cash'])}")
print(f"Max weight: {max(p[6] for p in POSITIONS if p[5] != 'Cash'):.0%}")
print(f"Sum >5%: {sum(p[6] for p in POSITIONS if p[5] != 'Cash' and p[6] > 0.05):.0%}")
